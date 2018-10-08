import sys
import heapq
import os
import shutil
from copy import deepcopy
from pprint import pprint

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.chart import LineChart, Reference

import config


# shortest_path procedure is from http://code.activestate.com/recipes/119466-dijkstras-algorithm-for-shortest-paths/
# Chris Laffra - best one
def shortest_path(graph, start, end):
    queue = [(0, start, [])]
    seen = set()
    while True:
        (cost, v, path) = heapq.heappop(queue)
        if v not in seen:
            path = path + [v]
            seen.add(v)
            if v == end:
                return cost, path
            for (next, c) in graph[v].items():
                heapq.heappush(queue, (cost + c, next, path))

def calculate_lens_and_paths(graph):
    lens = {}
    paths = {}
    for i in nodes_12:
        lens[i] = {}
        paths[i] = {}
        for j in nodes_12:
            lenght, path = shortest_path(graph, i, j)
            lens[i][j] = lenght
            if len(path) == 1:
                path = []
            paths[i][j] = path
    return lens, paths

def calculate_Dij(lens, flows, nodes, correction_coefs=None):
    Dij = {}

    for i in nodes:
        Dij[i] = {}
        for j in nodes:
            HPj = flows[j]['absorption']

            if i == j:
                Cij = 0.01 + 0.01 * 2   # 2 - остання цифра заліковки
            else:
                Cij = lens[i][j]**-1

            if correction_coefs:
                result = round(HPj * Cij * correction_coefs[j], 2)
            else:
                result = round(HPj * Cij, 2)

            Dij[i][j] = result

    return Dij

def calculate_matrix_column(j, matrix):
    result = 0
    for i in nodes:
        result = result + matrix[i][j]
    return result

def calculate_matrix_row(i, matrix):
    result = 0
    for j in nodes:
        result = result + matrix[i][j]
    return result

def calculate_correspondences(lens, flows, nodes, Dij):
    correspondences = {}

    for i in nodes:
        correspondences[i] = {}
        for j in nodes:
            HOi = flows[i]['creation']
            top = Dij[i][j]
            bottom = calculate_matrix_row(i, Dij)

            result = HOi * (top / bottom)
            correspondences[i][j] = round(result, 0)

    return correspondences

# dela_j - difference between given and calculated flow in % (must be as low as possible; delta_j < 5%*)
def calculate_delta_j_and_correction_coefs(flows, correspondences, calc_correction_coefs=True):
    delta_j = {}
    if calc_correction_coefs:
        correction_coefs = {}

    for j in nodes:
        # calculated by me
        calc_HPj = calculate_matrix_column(j, correspondences)
        # given absorption flow
        start_HPj = flows[j]['absorption']
        
        result = ((calc_HPj - start_HPj) / start_HPj) * 100
        delta_j[j] = round(result, 2)

        if calc_correction_coefs:
            k = start_HPj / calc_HPj
            correction_coefs[j] = k

    if calc_correction_coefs:
        return delta_j, correction_coefs
    return delta_j

def test_calculations(correspondences, flows):
    rows = []
    columns = []

    # probably can have only one loop and pass `i` in both functions
    for i in nodes:
        row = calculate_matrix_row(i, correspondences)
        rows.append(row)

    for j in nodes:
        column = calculate_matrix_column(j, correspondences)
        columns.append(column)
    
    result = zip(rows, columns)
    for i, values in enumerate(result):
        print('{}. creation: {}; absorption: {}'.format(i+1, values[0], values[1]))

def sumbit_rows_and_columns(correspondences):
    row = 0
    column = 0
    for i in nodes:
        for j in nodes:
            row = row + calculate_matrix_row(i, correspondences)
            column = column + calculate_matrix_column(j, correspondences)
    row = round(row, 2)
    column = round(column, 2)
    print('sumbit rows: {}'.format(row))
    print('sumbit columns: {}'.format(column))

def get_xls_styles():
    font = Font(bold=True, italic=True)

    border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin',color='FF000000')
    )
    return font, border

def write_table10x10(ws, name, data):
    # writing name of table
    ws.cell(row=1, column=1, value=name)

    # set up styles to border
    font, border = get_xls_styles()

    # writing column names and row indexes
    for column in nodes:
        column = int(column)
        
        # column name
        # +1 in order to have place for row indexes
        cell = ws.cell(row=2, column=column+1, value=column)
        cell.border = border
        cell.font = font

        # row index
        # +2 in order to save space for title and column names
        row = column
        cell = ws.cell(row=row+2, column=1, value=row)
        cell.border = border
        cell.font = font

    # writing data into sheet
    for i in nodes:
        for j in nodes:
            row = int(i) + 2    # to compensate header
            column = int(j) + 1     # to compensate row indexes
            value = data[i].get(j, '')

            if type(value) == list:
                value = '>'.join(value) or '-'

            cell = ws.cell(row=row, column=column, value=value)
            cell.border = border

def write_table1x10(ws, name, data):
    # writing name of table
    ws.cell(row=1, column=1, value=name)

    # set up styles to border
    font, border = get_xls_styles()

    # writing column names and row indexes
    for column in nodes:
        column = int(column)
        
        # column name
        # +1 in order to have place for row indexes
        cell = ws.cell(row=2, column=column+1, value=column)
        cell.border = border
        cell.font = font

        # row index
        # +2 in order to save space for title and column names
        if int(column) < 2:
            row = column
            cell = ws.cell(row=row+2, column=1, value=row)
            cell.border = border
            cell.font = font

    # writing data into prepared sheet
    for j in nodes:
        value = data[j]
        cell = ws.cell(row=3, column=int(j)+1, value=round(value, 2))
        cell.border = border

def write_table12x12(ws, name, data):
    # writing name of table
    ws.cell(row=1, column=1, value=name)

    # set up styles to border
    font, border = get_xls_styles()

    # writing column names and row indexes
    for column in nodes_12:
        column = int(column)
        
        # column name
        # +1 in order to have place for row indexes
        cell = ws.cell(row=2, column=column+1, value=column)
        cell.border = border
        cell.font = font

        # row index
        # +2 in order to save space for title and column names
        row = column
        cell = ws.cell(row=row+2, column=1, value=row)
        cell.border = border
        cell.font = font

    # writing data into sheet
    for i in nodes_12:
        for j in nodes_12:
            row = int(i) + 2    # to compensate header
            column = int(j) + 1     # to compensate row indexes

            value = data[i].get(j, '') or ''

            if type(value) == list:
                value = '>'.join(value) or '-'

            cell = ws.cell(row=row, column=column, value=value)
            cell.border = border

def write2excel(database, filename):
    wb = Workbook()

    # writing tables 10x10
    for name, data in database['10x10'].items():
        ws = wb.create_sheet(name)
        write_table10x10(ws, name, data)

    for name, data in database['1x10'].items():
        ws = wb.create_sheet(name)
        write_table1x10(ws, name, data)

    for name, data in database['single'].items():
        ws = wb.create_sheet(name)
        ws.cell(row=1, column=1, value=name)
        ws.cell(row=2, column=1, value=data)

    for name, data in database['12x12'].items():
        ws = wb.create_sheet(name)
        write_table12x12(ws, name, data)

    # change width of columns in order to properly see large data
    ws = wb.get_sheet_by_name('Шляхи')
    for column in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[column].width = 15

    to_remove = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(to_remove)

    wb.save(filename)


# probably can have one function insted two below
def calculate_transport_work(correspondences, lens):
    transport_work = {}

    for i in nodes:
        transport_work[i] = {}
        for j in nodes:
            transport_work[i][j] = round(correspondences[i][j] * lens[i][j], 2)

    return transport_work

def calculate_min_transport_work(transport_work):
    min_transport_work = 0

    for i in nodes:
        for j in nodes:
            min_transport_work += transport_work[i][j]

    return min_transport_work

def get_arcs(path):
    """
    Argumets:
        path (list of integers) - path from `i` to `j` described by nodes (dots)
    
    Returns:
        arcs (list of 2-length tuples) - path from `i` to `j` described by arcs (lines)

    Example:
        path = [1, 2, 3, 4]
        arcs = [(1, 2), (2, 3), (3, 4)]
    """
    arcs = []
    for i in range(len(path)):
        if i == len(path) - 1:
            break
        arc = (path[i], path[i + 1])
        arcs.append(arc)
    return arcs

# Notes
# correspondences has 10 nodes, because it is impossible to calculate 12,
# because data given only for 10
def calculate_passenger_flow(graph, paths, correspondences):
    passenger_flows = {}

    for i in graph:
        passenger_flows[i] = {}
        for j in graph[i]:
            # In this place frist pair of loops iterates through nodes (12 included)
            # second pair of loops iterates through 10
            # check whether here is a bug or not ...

            # AFTER CHECK:
            # `bug` is too loud
            # it's more like defect (describet before this function declaration)
            # actually the pas_flow number incomplete because we don't have
            # correspondences values for 10+ nodes but there are paths for that.

            # maybe it don't even cause an mistake on calculations

            #if int(i) > 10 or int(j) > 10:
            #    continue

            pas_flow = 0

            for m in nodes:
                for n in nodes:

                    if (i, j) in get_arcs(paths[m][n]):
                        try:
                            pas_flow += correspondences[m][n]
                        except:
                            continue

            passenger_flows[i][j] = round(pas_flow, 0)

    return passenger_flows

def check_straight_and_reverse(passenger_flows):
    straight_flow = 0
    reverse_flow = 0

    for i in passenger_flows:
        for j in passenger_flows[i]:

            if int(i) < int(j):
                straight_flow += passenger_flows[i][j]
            else:
                reverse_flow += passenger_flows[i][j]

    return straight_flow, reverse_flow

def calculate_general_pas_flow(passenger_flows, lens):
    general_pas_flow = 0

    for i in passenger_flows:
        for j in passenger_flows[i]:

            if int(i) > int(j):
                continue

            result = ( passenger_flows[i][j] + passenger_flows[j][i] ) * lens[i][j]
            general_pas_flow = general_pas_flow + result

    return round(general_pas_flow, 1)

def collect_values(d):
    # or just use [value for values in [d.values() for d in passenger_flows.values()] for value in list(values)]
    if type(d) != dict:
        return None

    values = []

    for value in d.values():

        if type(value) == dict:
            values += collect_values(value)
        else:
            values.append(value)

    return values

def make_recommendations(passenger_flows):
    all_values = collect_values(passenger_flows)
    high = max(all_values)
    low = min(all_values)

    delta = (high - low) / 3

    recommendations = {}

    for i in passenger_flows:
        recommendations[i] = {}
        for j in passenger_flows[i]:

            if low <= passenger_flows[i][j] <= low+delta:
                recommendations[i][j] = 1

            elif low+delta <= passenger_flows[i][j] <= high-delta:
                recommendations[i][j] = 2

            elif high-delta <= passenger_flows[i][j] <= high:
                recommendations[i][j] = 3

            else:
                recommendations[i][j] = 0

    return recommendations

def check_routes(routes):
    # init dict with all zeros in order to be able increment them further
    connections = { i: {} for i in nodes_12 for j in nodes_12}
    [ connections[i].update({j: 0}) for i in nodes_12 for j in nodes_12 ]

    for i in nodes_12:
        for j in nodes_12:

            for name, route in routes.items():
                if i != j and (i in route and j in route):
                    connections[i][j] += 1

    return connections

def len_route(route, lens):
    lenght = 0
    arcs = get_arcs(route)

    for i, j in arcs:
        lenght = lenght + lens[i][j]

    return round(lenght, 2)

def calculate_routes_lens(routes, lens):
    routes_lens = {}

    for name, route in routes.items():
        routes_lens[name] = len_route(route, lens)

    return routes_lens

def calculate_tranship_correspondence(connections, correspondences):
    tranship_correspondence = 0

    for i in connections:
        for j in connections[i]:

            if not connections[i][j] and i != j:
                # Here we face with a problem, that input data doesn't have creation and absorption for nodes 11 and 12
                # That's why I'm put try/except here.
                # We can assume those to 0, as one of the solutions. This mean, that those nodes will be just trasnit (just to ride through).
                try:
                    tranship_correspondence += correspondences[i][j]
                except:
                    pass

    return tranship_correspondence

def calculate_tranship_coef(correspondences, tranship_correspondence):
    correspondences_sum = sum(collect_values(correspondences))

    tranship_coef = (correspondences_sum + tranship_correspondence) / correspondences_sum

    return tranship_coef

def find_route_max_pas_flow(route, passenger_flows):
    route_max_pas_flow = 0

    for i, j in get_arcs(route):
        if passenger_flows[i][j] > route_max_pas_flow:
            route_max_pas_flow = passenger_flows[i][j]

        if passenger_flows[j][i] > route_max_pas_flow:
            route_max_pas_flow = passenger_flows[j][i]

    return route_max_pas_flow

def calc_top(route_path, passenger_flows, lens):
    results = []

    for i, j in get_arcs(route_path):
        results.extend([passenger_flows[i][j] * lens[i][j], passenger_flows[j][i] * lens[j][i]])

    return round(sum(results), 2)

# maybe write route builder (evolutional algorithm)
def calculate_routes_efficiency(lens, routes, routes_lens, passenger_flows):
    routes_efficiency_coefs = {}

    for route_num, route_path in routes.items():
        top = calc_top(route_path, passenger_flows, lens)
        bottom = 2 * find_route_max_pas_flow(route_path, passenger_flows) * routes_lens[route_num]

        #print(f'#{route_num}: {top} / {bottom}')

        routes_efficiency_coefs[route_num] = round(top / bottom, 2)

    pprint(routes_efficiency_coefs)
    average = str( sum(routes_efficiency_coefs.values()) / len(routes_efficiency_coefs) )
    routes_efficiency_coefs['average'] = average
    print('routes efficiency average: ', average)
    return routes_efficiency_coefs

def redistribute_correspondences(correspondences, connections):
    redistributed_correspondences = {}

    for i in correspondences:
        redistributed_correspondences[i] = {}
        for j in correspondences[i]:
            try:
                redistributed_correspondences[i][j] = correspondences[i][j] / connections[i][j]
            except:
                redistributed_correspondences[i][j] = 'div/zero'

    return redistributed_correspondences


def main(create_xls=True):

    # main data storage
    mds = {}
    mds['10x10'] = {}
    mds['1x10'] = {}
    mds['single'] = {}
    mds['12x12'] = {}

    lens, paths = calculate_lens_and_paths(graph)
    mds['12x12']['Найкоротшi вiдстанi'] = lens
    mds['12x12']['Шляхи'] = paths

    Dij = calculate_Dij(lens, flows, nodes)
    correspondences = calculate_correspondences(lens, flows, nodes, Dij)
    mds['10x10']['Функція тяжіння між вузлами'] = Dij
    mds['10x10']['Кореспонденцiї'] = correspondences

    #test_correspondences_calculations(flows, correspondences)
    before_delta_j, correction_coefs = calculate_delta_j_and_correction_coefs(flows, correspondences)
    mds['1x10']['Δj before'] = before_delta_j      # before correction*
    mds['1x10']['Поправочні коефіцієнти'] = correction_coefs

    corrected_Dij = calculate_Dij(lens, flows, nodes, correction_coefs)
    corrected_correspondences = calculate_correspondences(lens, flows, nodes, corrected_Dij)
    mds['10x10']['Функц.тяж.між.вузл(скорегов)'] = corrected_Dij
    mds['10x10']['Кореспонденцiї (скорегована)'] = corrected_correspondences

    after_delta_j = calculate_delta_j_and_correction_coefs(flows, corrected_correspondences, calc_correction_coefs=False)
    mds['1x10']['Δj after'] = after_delta_j

    transport_work = calculate_transport_work(corrected_correspondences, lens)
    min_transport_work = calculate_min_transport_work(transport_work)
    mds['10x10']['Транспортна робота'] = transport_work
    mds['single']['Мін. транспортна робота'] = min_transport_work

    passenger_flows = calculate_passenger_flow(graph, paths, corrected_correspondences)
    straight_flow, reverse_flow = check_straight_and_reverse(passenger_flows)
    mds['12x12']['Пасажиропотік'] = passenger_flows
    mds['single']['Сума пас. потоків'] = f'Прямий напрям: {straight_flow}; Зворотній напрям: {reverse_flow}'.replace('.', ',')

    general_pas_flow = calculate_general_pas_flow(passenger_flows, lens)    # fix bug: lens instead graph (in new-project branch)
    mds['single']['Заг. пас. потік'] = f'Загальний пасажиропотік: {general_pas_flow}'.replace('.', ',')

    recommendations = make_recommendations(passenger_flows)
    mds['12x12']['Рекомендації к-сті маршрутів'] = recommendations

    if routes:
        connections = check_routes(routes)
        mds['12x12']['Матриця зв\'язків'] = connections

        routes_lens = calculate_routes_lens(routes, lens)
        mds['single']['Довжини маршрутів'] = '\n'.join([f'{name}: {lenght}; ' for name, lenght in routes_lens.items()])

        tranship_correspondence = calculate_tranship_correspondence(connections, corrected_correspondences)
        tranship_coef = calculate_tranship_coef(corrected_correspondences, tranship_correspondence)
        mds['single']['Коефіцієнт пересаджуваності'] = tranship_coef

        routes_efficiency_coefs = calculate_routes_efficiency(lens, routes, routes_lens, passenger_flows)
        mds['single']['Коеф. ефективності маршрутів'] = '\n'.join([f'{name}: {coef}; ' for name, coef in routes_efficiency_coefs.items()])

        redistributed_correspondences = redistribute_correspondences(corrected_correspondences, connections)
        mds['10x10']['Перерозподілені Кореспонденцiї'] = redistributed_correspondences

        redistributed_passenger_flows = calculate_passenger_flow(graph, paths, redistributed_correspondences)
        mds['10x10']['Перерозподілені пас. потоки'] = redistributed_passenger_flows

    else:
        print("Info: Routes doesn't setup. Please, set them up and run program again in order to carry out related calculations.")

    # test cases

    #sumbit_rows_and_columns(correspondences)
    #test_calculations(correspondences, flows)

    if create_xls:
        filename = 'Розрахунки-{}.xlsx'.format(username)

        write2excel(mds, filename)

        if not os.path.exists('Results'):
            os.makedirs('Results')

        source = os.path.realpath(filename)
        destination = os.path.realpath('Results/{}'.format(filename))

        shutil.move(source, destination)    # move result file in "Results" directory

    return mds

def test_start_flows_equals(flows):
    creation = 0
    absorption = 0

    for n in nodes:
        creation += flows[n]['creation']
        absorption += flows[n]['absorption']

    return creation == absorption

def test_correspondences_calculations(flows, correspondences):
    for i in nodes:
        HOi = flows[i]['creation']
        for j in nodes:
            HPj = flows[j]['absorption']
            print('HOi: {} = {}'.format(HOi, calculate_matrix_row(i, correspondences)))
            print('HPj: {} = {}'.format(HPj, calculate_matrix_column(j, correspondences)))


if __name__ == '__main__':
    print('Welcome!')
    print('Thank you for using our service. Enjoy.')

    if len(sys.argv) != 2:
        print(f'usage: {__file__} username\nexample: {__file__} Роман-Худобей')
        sys.exit(0)

    username = sys.argv[1]

    graphs = {'Роман-Худобей': config.MY_GRAPH,
              'Віталій-Стахів': config.STAHIV_GRAPH}

    flows = {'Роман-Худобей': config.MY_FLOWS,
             'Віталій-Стахів': config.STAHIV_FLOWS}

    routes = {'Роман-Худобей': config.MY_ROUTES}

    #################################################################################
    #                               I M P O R T A N T                               #
    #                                                                               #
    # graph = {'1': ['6', '11', '2']}                                               #
    # lens = {'1': {'6': 1.5, '11': 2, '2': 1.2}}                                   #
    # shortest_paths = {'1': {'6': ['1', '6'], '11': ['1', '11'], '2': ['1', '2']}} #
    #                                                                               #
    #################################################################################

    graph = graphs.get(username)
    flows = flows.get(username)     # overwriting in case we don't need previous variable anymore
    routes = routes.get(username)

    if graph is None or flows is None:
        print("There's no data to start with.\nCheck if you typed username right or you forgot to register a user in config file.")
        sys.exit(0)

    nodes = config.NODES
    nodes_12 = config.NODES_12

    assert test_start_flows_equals(flows), "Please, check flows entered. Creation not equal absorption"

    main()

    print('Done. Congrats! You just saved few lives of your life.')