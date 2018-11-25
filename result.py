import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart, ScatterChart, Series
import xlsxwriter as xls    # TODO: Rewrite in xlsxwriter


class Result(object):
    RESULT_TYPES = ('12x12', '10x10', '1x10', 'route', 'single', 'pandas', 'rows')

    def __init__(self, rtype, title='Title', result={}, **kwargs):
        assert rtype in self.RESULT_TYPES, f'Type {rtype} is not valid'
        self.rtype = rtype  # TODO: type(<Result object at 0x063278ff>) if possible
        assert title, 'Title cannot be empty'
        self.title = str(title)
        self.data = result
        self.extra = kwargs


class ExcelResultsWriter(object):
    font = Font(bold=True, italic=True)

    border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
    )

    def __init__(self, results):
        self.results = results
        self.rtype_to_method = {
            '12x12': self.__write_table12x12,
            '10x10': self.__write_table10x10,
            '1x10': self.__write_table1x10,
            'route': self.__write_route,
            'single': self.__write_single,
            'pandas': self.__write_by_pandas,
            'rows': self.__write_rows
        }

    def write2excel(self, filename):
        # wb = workbook
        # ws = worksheet
        wb = Workbook()

        for result in self.results.values():
            ws = wb.create_sheet(result.title)
            self.rtype_to_method[result.rtype](ws, result.title, result.data, **result.extra)

        # change width of columns in order to properly see large data
        ws = wb['Шляхи']
        for column in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[column].width = 15

        wb.remove(wb['Sheet'])  # remove default

        wb.save(filename)

    def __write_table12x12(self, ws, name, data):
        # writing table title
        ws.cell(row=1, column=1, value=name)

        # writing row and column indexes/names
        for column in range(1, 13):
            # column name
            # +1 in order to have place for row indexes
            cell = ws.cell(row=2, column=column+1, value=column)
            cell.border = self.border
            cell.font = self.font

            # row index
            # +2 in order to save space for title and column names
            row = column
            cell = ws.cell(row=row+2, column=1, value=row)
            cell.border = self.border
            cell.font = self.font

        # writing data into sheet
        for i in (str(i) for i in range(1, 13)):
            for j in (str(i) for i in range(1, 13)):
                row = int(i) + 2  # to compensate header
                column = int(j) + 1  # to compensate row indexes

                value = data[i].get(j, '')

                if type(value) == list:     # path
                    value = '>'.join(value) or '-'

                cell = ws.cell(row=row, column=column, value=value)
                cell.border = self.border

    def __write_table10x10(self, ws, name, data):
        # writing table title
        ws.cell(row=1, column=1, value=name)

        # writing row and column indexes/names
        for column in range(1, 11):
            # column name
            # +1 in order to have place for row indexes
            cell = ws.cell(row=2, column=column+1, value=column)
            cell.border = self.border
            cell.font = self.font

            # row index
            # +2 in order to save space for title and column names
            row = column
            cell = ws.cell(row=row+2, column=1, value=row)
            cell.border = self.border
            cell.font = self.font

        # writing data into sheet
        for i in (str(i) for i in range(1, 11)):
            for j in (str(i) for i in range(1, 11)):
                row = int(i) + 2  # to compensate header
                column = int(j) + 1  # to compensate row indexes

                value = data[i].get(j, '')

                if type(value) == list:
                    value = '>'.join(value) or '-'

                cell = ws.cell(row=row, column=column, value=value)
                cell.border = self.border

    def __write_table1x10(self, ws, name, data):
        # writing table title
        ws.cell(row=1, column=1, value=name)

        # writing row and column indexes/names
        for column in range(1, 11):
            # column name
            # +1 in order to have place for row indexes
            cell = ws.cell(row=2, column=column+1, value=column)
            cell.border = self.border
            cell.font = self.font

            # row index
            # +2 in order to save space for title and column names
            if column < 2:
                row = column
                cell = ws.cell(row=row+2, column=1, value=row)
                cell.border = self.border
                cell.font = self.font

        # writing data into prepared sheet
        for j in (str(i) for i in range(1, 11)):
            value = data[j]
            cell = ws.cell(row=3, column=int(j) + 1, value=value)  # round here to have more accurate results
            cell.border = self.border

    def __write_route(self, ws, name, route_obj, **kwargs):
        LAST_ROW = 0
        ws.append([name, str(route_obj)])
        LAST_ROW += 1
        ws.append(['Ефективність', route_obj.efficiency()])
        LAST_ROW += 1
        ws.append([])
        LAST_ROW += 1
        ws.append(['Напрям перегону'] + ['-'.join(arc) for arc in route_obj.arcs])
        LAST_ROW += 1
        ws.append(['Прямий напрям'] + [route_obj.passenger_flow[i][j] for i, j in route_obj.arcs])
        LAST_ROW += 1
        ws.append(['Зворотній напрям'] + [route_obj.passenger_flow[j][i] for i, j in route_obj.arcs])
        LAST_ROW += 1
        ws.append([])
        LAST_ROW += 1

        if kwargs.get('include_missed_flow_chart'):
            ws.append(['Дані до епюри неврахованого пасажиропотоку (До рисунку 6.7)'])
            LAST_ROW += 1
            for row in route_obj.missed_flow_table_part():
                ws.append(row)
                LAST_ROW += 1

            missed_flow_chart = ScatterChart()
            missed_flow_chart.style = 10
            missed_flow_chart.title = 'Епюра неврахованого пасажиропотоку'
            missed_flow_chart.y_axis.title = 'Пасажиропотік'
            missed_flow_chart.x_axis.title = 'Перегон маршруту'

            xvalues = Reference(ws, min_col=2, min_row=9, max_col=len(route_obj.arcs)+1, max_row=9)    # x axis labels

            values1 = Reference(ws, min_col=1, min_row=10, max_col=len(route_obj.arcs)+1, max_row=10)   # data for chart
            values2 = Reference(ws, min_col=1, min_row=11, max_col=len(route_obj.arcs)+1, max_row=11)

            s1 = Series(values1, xvalues, title_from_data=True)
            s2 = Series(values2, xvalues, title_from_data=True)

            missed_flow_chart.series.append(s1)
            missed_flow_chart.series.append(s2)

            missed_flow_chart.shape = 4

            ws.append([])
            LAST_ROW += 1

        for row in route_obj.slice_from_redistributed_correspondences(heading=f"{'Доповнена матриця міжрайонних кореспонденцій для маршруту' if kwargs.get('include_missed_flow_chart') else 'Матриця міжрайонних кореспонденцій для маршруту'}", include_missed_flow=kwargs.get('include_missed_flow_chart')):
            ws.append(row)
            LAST_ROW += 1

        ws.append([])
        LAST_ROW += 1

        chart = ScatterChart()
        chart.style = 10
        chart.title = 'Епюра пасажиропотоку'
        chart.y_axis.title = 'Пасажиропотік'
        chart.x_axis.title = 'Перегон маршруту'

        xvalues = Reference(ws, min_col=2, min_row=4, max_col=len(route_obj.arcs)+1, max_row=4)    # x axis labels

        values1 = Reference(ws, min_col=1, min_row=5, max_col=len(route_obj.arcs)+1, max_row=5)   # data for chart
        values2 = Reference(ws, min_col=1, min_row=6, max_col=len(route_obj.arcs)+1, max_row=6)

        s1 = Series(values1, xvalues, title_from_data=True)
        s2 = Series(values2, xvalues, title_from_data=True)

        chart.series.append(s1)
        chart.series.append(s2)

        chart.shape = 4

        if kwargs.get('include_missed_flow_chart'):
            ws.add_chart(missed_flow_chart, f'A{LAST_ROW+1}')
            LAST_ROW += 15

        ws.add_chart(chart, f"A{LAST_ROW+1}")
        LAST_ROW += 15

        # self.__write_route2(ws, name, route_obj)

    def __write_route2(self, ws, name, route_obj):
        wb = xls.Workbook(f"{name}.xlsx")
        ws = wb.add_worksheet()

        # add route path

        ws.write_row('A1', ['Напрям перегону'] + ['-'.join(arc) for arc in route_obj.arcs])
        ws.write_row('A2', ['Прямий напрям'] + [route_obj.passenger_flow[i][j] for i, j in route_obj.arcs])
        ws.write_row('A3', ['Зворотній напрям'] + [route_obj.passenger_flow[j][i] for i, j in route_obj.arcs])

        # chart = wb.add_chart({'type': 'column'})
        chart = wb.add_chart({'type': 'line'})

        chart.add_series({
            'name': '=Sheet1!$A$2',
            'categories': '=Sheet1!$B$1:$F$1',
            'values': '=Sheet1!$B$2:$F$2'
        })
        chart.add_series({
            'name': '=Sheet1!$A$3',
            'categories': '=Sheet1!$B$1:$F$1',
            'values': '=Sheet1!$B$3:$F$3'
        })

        chart.set_style(10)
        ws.insert_chart('D2', chart)

        wb.close()

    def __write_by_pandas(self, ws, name, data, **kwargs):
        pd.set_option('display.max_colwidth', -1)
        # writing table title
        ws.append([name])

        if type(data) != pd.DataFrame:
            table = pd.DataFrame.from_dict(data)
            table.index = ('' for _ in range(len(data[list(data.keys())[0]])))  # len of column/row (table consistency assumed)

        if kwargs.get('transpose'):
            table = table.T

        rows = list(row.split(';') for row in table.to_csv(sep=';').split('\n'))

        if kwargs.get('transpose'):
            rows[0] = [''] + rows[0]    # add column offset to column indexes (compensate row indexes)

        for row in rows:
            row = [int(value) if value.isdigit() else value for value in row if value]

            ws.append(row)

        if kwargs.get('chart_config'):
            chart_types = {
                'column': BarChart,
                'line': ScatterChart
            }
            chart_config = kwargs.get('chart_config')
            chart = chart_types[chart_config['type']]()
            chart.style = 10
            chart.title = name
            chart.y_axis.title = chart_config['y_title']
            chart.x_axis.title = chart_config['x_title']

            xvalues = Reference(ws, min_col=1, min_row=3, max_col=len(rows[0])+1, max_row=3)  # x axis labels
            values1 = Reference(ws, min_col=1, min_row=4, max_col=len(rows[0])+1, max_row=len(rows)+1)  # data for chart
            s1 = Series(values1, xvalues, title_from_data=True)
            chart.series.append(s1)

            # cats = Reference(ws, min_col=1, min_row=3, max_col=len(rows[0]), max_row=3)
            # data = Reference(ws, min_col=1, min_row=4, max_col=len(rows[0]), max_row=len(rows))
            #
            # chart.add_data(data)
            # chart.set_categories(cats)

            chart.shape = 4
            ws.add_chart(chart, f'A{len(rows)+2}')

    def __write_single(self, ws, name, data):
        ws.cell(row=1, column=1, value=name)
        ws.cell(row=2, column=1, value=data)

    def __write_rows(self, ws, name, data):
        ws.append([name])
        for row in data: ws.append(row)


class XlsxResultsWriter(object):

    def __init__(self, results):
        self.results = results
        self.data_type_to_method = {
            'table': self.__write_table
        }

    def write2excel(self, filename):
        wb = xls.Workbook(filename)

        for result in self.results.values():
            ws = wb.add_worksheet(name=result.title)
            self.data_type_to_method[result.data_type](ws, result.title, result.data)

        wb.close()

    def __write_table(self, ws, name, data):
        ws.write_row()
